import os
import streamlit as st
import pandas as pd
import requests
import base64
import hashlib
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font


GITHUB_TOKEN = st.secrets["GITHUB_TOKEN"]  # 在 Streamlit Cloud 用 secrets
REPO_NAME = "TTTriste06/semi"
BRANCH = "main"

CONFIG = {
    "output_file": f"运营数据订单-在制-库存汇总报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    "selected_month": None,
    "pivot_config": {
        "赛卓-未交订单.xlsx": {
            "index": ["晶圆品名", "规格", "品名"],
            "columns": "预交货日",
            "values": ["订单数量", "未交订单数量"],
            "aggfunc": "sum",
            "date_format": "%Y-%m"
        },
        "赛卓-成品在制.xlsx": {
            "index": ["工作中心", "封装形式", "晶圆型号", "产品规格", "产品品名"],
            "columns": "预计完工日期",
            "values": ["未交"],
            "aggfunc": "sum",
            "date_format": "%Y-%m"
        },
        "赛卓-CP在制.xlsx": {
            "index": ["晶圆型号", "产品品名"],
            "columns": "预计完工日期",
            "values": ["未交"],
            "aggfunc": "sum",
            "date_format": "%Y-%m"
        },
        "赛卓-成品库存.xlsx": {
            "index": ["WAFER品名", "规格", "品名"],
            "columns": "仓库名称",
            "values": ["数量"],
            "aggfunc": "sum"
        },
        "赛卓-晶圆库存.xlsx": {
            "index": ["WAFER品名", "规格"],
            "columns": "仓库名称",
            "values": ["数量"],
            "aggfunc": "sum"
        }
    }
}

def upload_to_github(file, path_in_repo, commit_message):
    api_url = f"https://api.github.com/repos/{REPO_NAME}/contents/{path_in_repo}"
    
    file.seek(0)  # 确保指针在开头
    file_content = file.read()
    encoded_content = base64.b64encode(file_content).decode('utf-8')

    # 先获取文件 SHA（如果存在）
    response = requests.get(api_url, headers={"Authorization": f"token {GITHUB_TOKEN}"})
    if response.status_code == 200:
        sha = response.json()['sha']
    else:
        sha = None

    # 构造 payload
    payload = {
        "message": commit_message,
        "content": encoded_content,
        "branch": BRANCH
    }
    if sha:
        payload["sha"] = sha

    # 上传文件
    response = requests.put(api_url, json=payload, headers={"Authorization": f"token {GITHUB_TOKEN}"})

    # 结果反馈
    if response.status_code in [200, 201]:
        st.success(f"{path_in_repo} 上传成功！")
    else:
        st.error(f"上传失败: {response.json()}")


def preprocess_mapping_file(df):
    # 只取前6列
    df = df.iloc[:, :6]
    # 重命名列
    df.columns = ['旧规格', '旧品名', '旧晶圆品名', '新规格', '新品名', '新晶圆品名']
    return df


def download_mapping_from_github(path_in_repo):
    api_url = f"https://api.github.com/repos/{REPO_NAME}/contents/{path_in_repo}"
    response = requests.get(api_url, headers={
        "Authorization": f"token {GITHUB_TOKEN}"
    })
    if response.status_code == 200:
        content = base64.b64decode(response.json()['content'])
        df = pd.read_excel(pd.io.common.BytesIO(content))
        if df.shape[1] >= 6:
            df = preprocess_mapping_file(df)
        else:
            st.warning(f"mapping_file.xlsx 列数不足：只发现 {df.shape[1]} 列，需要至少 6 列")
            df = pd.DataFrame(columns=['旧规格', '旧品名', '旧晶圆品名', '新规格', '新品名', '新晶圆品名'])
        return df
    else:
        st.warning("GitHub 上找不到 mapping_file.xlsx，用默认表或请先上传")
        return pd.DataFrame(columns=['旧规格', '旧品名', '旧晶圆品名', '新规格', '新品名', '新晶圆品名'])

def download_excel_from_github(url, token=None):
    headers = {"Authorization": f"token {token}"} if token else {}
    response = requests.get(url, headers=headers)
    content_type = response.headers.get('Content-Type', '')

    # 检查文件是不是 Excel
    if 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' not in content_type:
        raise ValueError("下载的不是 Excel 文件，请检查 GitHub 链接或权限")

    return pd.read_excel(BytesIO(response.content))

def download_backup_file(file_name):
    api_url = f"https://api.github.com/repos/{REPO_NAME}/contents/{file_name}"
    headers = {"Authorization": f"token {GITHUB_TOKEN}"}
    response = requests.get(api_url, headers=headers)

    if response.status_code != 200:
        st.warning(f"⚠️ 无法下载 {file_name}，GitHub 返回码 {response.status_code}")
        return pd.DataFrame()  # 返回空 DataFrame，保证主程序不崩溃

    content = response.json().get('content')
    if not content:
        st.warning(f"⚠️ {file_name} 文件内容为空或解析失败")
        return pd.DataFrame()

    file_bytes = BytesIO(base64.b64decode(content))

    try:
        df = pd.read_excel(file_bytes)
    except Exception as e:
        st.warning(f"⚠️ {file_name} 解析 Excel 失败：{e}，将创建空 sheet。")
        return pd.DataFrame()

    return df
        
def process_date_column(df, date_col, date_format):
    if pd.api.types.is_numeric_dtype(df[date_col]):
        df[date_col] = df[date_col].apply(lambda x: datetime(1899, 12, 30) + timedelta(days=float(x)))
    else:
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    df[f'{date_col}_年月'] = df[date_col].dt.strftime(date_format)
    return df

def apply_mapping_and_merge(df, mapping_df):
    mapping_df = mapping_df.dropna()
    
    df = df.merge(
        mapping_df,
        how='left',
        left_on=['晶圆品名', '规格', '品名'],
        right_on=['旧晶圆品名', '旧规格', '旧品名']
    )
    
    df['晶圆品名'] = df['新晶圆品名'].combine_first(df['晶圆品名'])
    df['规格'] = df['新规格'].combine_first(df['规格'])
    df['品名'] = df['新品名'].combine_first(df['品名'])
    
    df.drop(columns=['旧晶圆品名', '旧规格', '旧品名', '新晶圆品名', '新规格', '新品名'], inplace=True)
    
    group_cols = [col for col in df.columns if col not in df.select_dtypes(include='number').columns]
    agg_cols = df.select_dtypes(include='number').columns.tolist()
    df_merged = df.groupby(group_cols, as_index=False)[agg_cols].sum()
    
    return df_merged

def create_pivot(df, config, filename, mapping_df=None):
    if 'date_format' in config:
        config = config.copy()
        config['columns'] = f"{config['columns']}_年月"
    pivoted = pd.pivot_table(df, index=config['index'], columns=config['columns'], values=config['values'],
                             aggfunc=config['aggfunc'], fill_value=0)
    pivoted.columns = [f"{col[0]}_{col[1]}" if isinstance(col, tuple) else col for col in pivoted.columns]
    pivoted = pivoted.reset_index()

    if mapping_df is not None and filename == "赛卓-未交订单.xlsx":
        pivoted = apply_mapping_and_merge(pivoted, mapping_df)

    if CONFIG['selected_month'] and filename == "赛卓-未交订单.xlsx":
        history_cols = [col for col in pivoted.columns if '_' in col and col.split('_')[-1][:4].isdigit() and col.split('_')[-1] < CONFIG['selected_month']]
        history_order_cols = [col for col in history_cols if '订单数量' in col and '未交订单数量' not in col]
        history_pending_cols = [col for col in history_cols if '未交订单数量' in col]
        if history_order_cols:
            pivoted['历史订单数量'] = pivoted[history_order_cols].sum(axis=1)
        if history_pending_cols:
            pivoted['历史未交订单数量'] = pivoted[history_pending_cols].sum(axis=1)
        pivoted.drop(columns=history_cols, inplace=True)
        fixed_cols = [col for col in pivoted.columns if col not in ['历史订单数量', '历史未交订单数量']]
        if '历史订单数量' in pivoted.columns:
            fixed_cols.insert(len(config['index']), '历史订单数量')
        if '历史未交订单数量' in pivoted.columns:
            fixed_cols.insert(len(config['index']) + 1, '历史未交订单数量')
        pivoted = pivoted[fixed_cols]

    return pivoted

def adjust_column_width(writer, sheet_name, df):
    worksheet = writer.sheets[sheet_name]
    for idx, col in enumerate(df.columns, 1):
        max_len = df[col].astype(str).str.len().max()
        header_len = len(str(col))
        width = max(max_len, header_len) * 1.2 + 5
        worksheet.column_dimensions[get_column_letter(idx)].width = min(width, 50)

def add_black_border(ws, row_count, col_count):
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=row_count, min_col=1, max_col=col_count):
        for cell in row:
            cell.border = border

def main():
    st.set_page_config(
        page_title='我是标题',
        page_icon=' ',
        layout='wide'
    )

    with st.sidebar:
        st.title("欢迎来到我的应用")
        st.markdown('---')
        st.markdown('这是它的特性：\n- feature 1\n- feature 2\n- feature 3')

    st.title('Excel 数据处理与汇总工具')
    selected_month = st.text_input('请输入截至月份（如 2025-03，可选）')
    CONFIG['selected_month'] = selected_month if selected_month else None

    uploaded_files = st.file_uploader('上传 Excel 文件（5个文件）', type=['xlsx'], accept_multiple_files=True)
    pred_file = st.file_uploader('上传预测文件', type=['xlsx'], key='pred_file')
    safety_file = st.file_uploader('上传安全库存文件', type=['xlsx'], key='safety_file')
    mapping_file = st.file_uploader('上传新旧料号文件', type=['xlsx'], key='mapping_file')

    # 加载 mapping_file DataFrame
    mapping_df = None
    if mapping_file:
        mapping_df = pd.read_excel(mapping_file)
        mapping_df = preprocess_mapping_file(mapping_df)

    if pred_file:
        upload_to_github(pred_file, "pred_file.xlsx", "上传预测文件")
    if safety_file:
        upload_to_github(safety_file, "safety_file.xlsx", "上传安全库存文件")
    if mapping_file:
        upload_to_github(mapping_file, "mapping_file.xlsx", "上传新旧料号文件")

    if st.button('提交并生成报告') and uploaded_files:
        with pd.ExcelWriter(CONFIG['output_file'], engine='openpyxl') as writer:
            # 用于存储未交订单的前三列数据
            unfulfilled_orders_summary = pd.DataFrame()
            df_safety = pd.DataFrame()

            for f in uploaded_files:
                filename = f.name
                if filename not in CONFIG['pivot_config']:
                    st.warning(f"跳过未配置的文件: {filename}")
                    continue

                df = pd.read_excel(f)
                config = CONFIG['pivot_config'][filename]

                if 'date_format' in config and config['columns'] in df.columns:
                    df = process_date_column(df, config['columns'], config['date_format'])

                pivoted = create_pivot(df, config, filename, mapping_df)
                sheet_name = filename[:30].rstrip('.xlsx')
                pivoted.to_excel(writer, sheet_name=sheet_name, index=False)
                adjust_column_width(writer, sheet_name, pivoted)

                # 保存未交订单的前三列（去重）
                if filename == "赛卓-未交订单.xlsx":
                    cols_to_copy = [col for col in pivoted.columns if col in ["晶圆品名", "规格", "品名"]]
                    unfulfilled_orders_summary = pivoted[cols_to_copy].drop_duplicates()
                    pending_pivoted = pivoted.copy()

            # 写入安全库存 sheet
            if safety_file:
                df_safety = pd.read_excel(safety_file)
            else:
                df_safety = download_backup_file("safety_file.xlsx")
            df_safety.to_excel(writer, sheet_name='赛卓-安全库存', index=False)
            adjust_column_width(writer, '赛卓-安全库存', df_safety)

            # 写入预测文件 sheet
            if pred_file:
                df_pred = pd.read_excel(pred_file)
            else:
                df_pred = download_backup_file("pred_file.xlsx")
            df_pred.to_excel(writer, sheet_name='赛卓-预测', index=False)
            adjust_column_width(writer, '赛卓-预测', df_pred)

            # 写入新旧料号文件 sheet
            # === 在处理成品在制之前，重新加载 mapping_file 全表 ===
            if mapping_file:
                df_full_mapping = pd.read_excel(mapping_file)
                
                # 设置列名（假设 Excel 里有9列，含封装厂、PC、半成品）
                df_full_mapping.columns = ['旧规格', '旧品名', '旧晶圆品名', '新规格', '新品名', '新晶圆品名', '封装厂', 'PC', '半成品']
            else:
                df_full_mapping = download_backup_file("mapping_file.xlsx")
                df_full_mapping.columns = ['旧规格', '旧品名', '旧晶圆品名', '新规格', '新品名', '新晶圆品名', '封装厂', 'PC', '半成品']
            
            # 写入新旧料号文件 sheet
            if mapping_file:
                df_mapping = pd.read_excel(mapping_file, header = 1)
            else:
                df_mapping = download_backup_file("mapping_file.xlsx")
           
            # 第3行开始写入数据（跳过第1、2行）
            df_mapping.to_excel(writer, sheet_name='赛卓-新旧料号', index=False, header=False, startrow=2)
            
            # 获取 worksheet
            ws = writer.book['赛卓-新旧料号']
            ws.delete_rows(0)

            # 写入第2行表头（DataFrame 的列名）
            for col_idx, col_name in enumerate(df_mapping.columns, start=1):
                ws.cell(row=2, column=col_idx, value=col_name)
                ws.cell(row=2, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=2, column=col_idx).font = Font(bold=True)
            
            # 写入第1行大标题（合并单元格）
            ws.merge_cells('A1:C1')
            ws['A1'] = '旧'
            ws.merge_cells('D1:F1')
            ws['D1'] = '新'
            
            # 设置第1行填充颜色、居中、加粗
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            
            for cell in ['A1', 'B1', 'C1']:
                ws[cell].fill = yellow_fill
            for cell in ['D1', 'E1', 'F1']:
                ws[cell].fill = green_fill
            
            for col in range(1, len(df_mapping.columns) + 1):
                ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=1, column=col).font = Font(bold=True)
            
            # 开启 Excel 筛选器（第2行是表头）
            from openpyxl.utils import get_column_letter
            last_col_letter = get_column_letter(len(df_mapping.columns))
            ws.auto_filter.ref = f"A2:{last_col_letter}2"

            # 定义新的列名
            new_column_names = ['旧规格', '旧品名', '旧晶圆品名', '新规格', '新品名', '新晶圆品名', '封装厂', 'PC', '半成品']
            
            # 直接重命名第二行每一列
            for col_idx, col_name in enumerate(new_column_names, start=1):
                ws.cell(row=2, column=col_idx, value=col_name)
                ws.cell(row=2, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=2, column=col_idx).font = Font(bold=True)
                
            # 自动调整列宽
            for idx, col in enumerate(ws.columns, 1):
                col_letter = get_column_letter(idx)
                max_length = 0
                for cell in col:
                    try:
                          if cell.value:
                            cell_len = sum(2 if ord(char) > 127 else 1 for char in str(cell.value))
                            max_length = max(max_length, cell_len)
                    except:
                           pass
                ws.column_dimensions[col_letter].width = max_length + 5



            # 写入汇总 sheet
            if not unfulfilled_orders_summary.empty:
                unfulfilled_orders_summary.to_excel(writer, sheet_name='汇总', index=False, startrow=1)
                adjust_column_width(writer, '汇总', unfulfilled_orders_summary)

                worksheet = writer.book['汇总']
                
                ###安全库存
                # 重命名安全库存列方便匹配
                df_safety.rename(columns={
                    'WaferID': '晶圆品名',
                    'OrderInformation': '规格',
                    'ProductionNO.': '品名'
                }, inplace=True)
                
                # 做一个标志列，表示是否被使用
                df_safety['已匹配'] = False
                
                # 合并安全库存数据到汇总 sheet（unfulfilled_orders_summary）
                summary_with_safety = unfulfilled_orders_summary.merge(
                    df_safety[['晶圆品名', '规格', '品名', ' InvWaf', ' InvPart']],
                    on=['晶圆品名', '规格', '品名'], 
                    how='left'
                )
                
                #更新已匹配标志
                matched_mask = df_safety.set_index(['晶圆品名', '规格', '品名']).index.isin(
                    summary_with_safety.dropna(subset=[' InvWaf', ' InvPart']).set_index(['晶圆品名', '规格', '品名']).index
                )
                df_safety.loc[matched_mask, '已匹配'] = True
                
                #写入汇总 sheet
                summary_with_safety.rename(columns={' InvWaf': 'InvWaf（片）', ' InvPart': 'InvPart'}, inplace=True)
                summary_with_safety.to_excel(writer, sheet_name='汇总', index=False, startrow=1)
                adjust_column_width(writer, '汇总', summary_with_safety)
                
                # 合并 D1:E1 写入表头
                worksheet = writer.book['汇总']
                worksheet.merge_cells('D1:E1')
                worksheet['D1'] = '安全库存'
                worksheet['D1'].alignment = Alignment(horizontal='center', vertical='center')
                worksheet['D2'] = 'InvWaf（片）'
                worksheet['D2'].alignment = Alignment(horizontal='center', vertical='center')
                worksheet['E2'] = 'InvPart'
                worksheet['E2'].alignment = Alignment(horizontal='center', vertical='center')
                
                # 标红未被使用的安全库存行
                safety_sheet = writer.book['赛卓-安全库存']
                for row_idx, used in enumerate(df_safety['已匹配'], start=2):  # Excel 从 1 开始，header 是第1行
                    if not used:
                        for col in range(1, len(df_safety.columns) + 1):
                            safety_sheet.cell(row=row_idx, column=col).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                ###未交订单
                if pending_pivoted is not None:
                    # 找出历史未交订单和未来每月未交订单列
                    pending_cols = [col for col in pending_pivoted.columns if '未交订单数量_' in col]
                    if '历史未交订单数量' in pending_pivoted.columns:
                        all_pending_cols = ['历史未交订单数量'] + pending_cols
                    else:
                        pending_pivoted['历史未交订单数量'] = 0
                        all_pending_cols = ['历史未交订单数量'] + pending_cols
                
                    # 计算总和
                    pending_pivoted['总未交订单'] = pending_pivoted[all_pending_cols].sum(axis=1)
                
                    # 整理顺序
                    pending_summary_cols = ['总未交订单'] + all_pending_cols
                    pending_summary_df = pending_pivoted[['晶圆品名', '规格', '品名'] + pending_summary_cols]
                
                    # 定位汇总 sheet
                    summary_sheet = writer.sheets['汇总']
                    # ✅ 前三列 + 安全库存两列 + 1
                    start_col = unfulfilled_orders_summary.shape[1] + 2 + 1  
                    end_col = start_col + len(pending_summary_cols) - 1
                
                    # ✅ 合并第一行所有新列（不覆盖安全库存）
                    summary_sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                    summary_sheet.cell(row=1, column=start_col, value='未交订单').alignment = Alignment(horizontal='center', vertical='center')
                
                    # ✅ 写入第二行标题
                    for idx, col in enumerate(pending_summary_cols, start=start_col):
                        summary_sheet.cell(row=2, column=idx, value=col).alignment = Alignment(horizontal='center', vertical='center')
                
                    # ✅ 写入第三行及以后数据
                    for row_idx, row in enumerate(pending_summary_df[pending_summary_cols].itertuples(index=False), start=3):
                        for col_idx, value in enumerate(row, start=start_col):
                            summary_sheet.cell(row=row_idx, column=col_idx, value=value)

                ### 预测
                # === 处理预测 sheet ===
                if not df_pred.empty:
                    df_pred.columns = df_pred.iloc[0]
                    df_pred = df_pred.drop([0,0]).reset_index(drop=True)
                
                    # 检查列是否存在
                    required_columns = ['晶圆品名', '产品型号', 'ProductionNO.', '合计数量', '合计金额']
                    if all(col in df_pred.columns for col in required_columns):
                        # 构造 key 进行匹配
                        pred_key = df_pred[['晶圆品名', '产品型号', 'ProductionNO.']].astype(str)
                        summary_key = unfulfilled_orders_summary[['晶圆品名', '规格', '品名']].astype(str)
                
                        # 在预测表中加匹配标志
                        df_pred['已匹配'] = False
                
                        # 定位汇总 sheet
                        summary_sheet = writer.sheets['汇总']
                        start_col = summary_sheet.max_column + 1  # 空白列的起点
                
                        # 合并第一行单元格写入“预测”
                        summary_sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col+1)
                        summary_sheet.cell(row=1, column=start_col, value='预测').alignment = Alignment(horizontal='center', vertical='center')
                        summary_sheet.cell(row=2, column=start_col, value='合计数量').alignment = Alignment(horizontal='center', vertical='center')
                        summary_sheet.cell(row=2, column=start_col+1, value='合计金额').alignment = Alignment(horizontal='center', vertical='center')
                
                        # 遍历汇总 sheet 的数据行（从第3行开始）
                        for row_idx in range(3, summary_sheet.max_row + 1):
                            summary_wf = summary_sheet.cell(row=row_idx, column=1).value
                            summary_spec = summary_sheet.cell(row=row_idx, column=2).value
                            summary_prod = summary_sheet.cell(row=row_idx, column=3).value
                
                            # 查找预测表匹配行
                            match = df_pred[
                                (df_pred['晶圆品名'].astype(str) == str(summary_wf)) &
                                (df_pred['产品型号'].astype(str) == str(summary_spec)) &
                                (df_pred['ProductionNO.'].astype(str) == str(summary_prod))
                            ]
                
                            if not match.empty:
                                qty = match['合计数量'].values[0]
                                amt = match['合计金额'].values[0]
                                summary_sheet.cell(row=row_idx, column=start_col, value=qty)
                                summary_sheet.cell(row=row_idx, column=start_col+1, value=amt)
                
                                # 标记预测表的行
                                df_pred.loc[match.index, '已匹配'] = True
                
                        # 在预测表中标红未匹配的行
                        pred_sheet = writer.book['赛卓-预测']
                        for row_idx, matched in enumerate(df_pred['已匹配'], start=3):  # 从第3行开始
                            if not matched:
                                for col_idx in range(1, len(df_pred.columns) + 1):
                                    pred_sheet.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                
                        # 给汇总表前两行加黑框
                        add_black_border(summary_sheet, 2, summary_sheet.max_column)
                        
                ###成品库存
                # === 从生成的赛卓-成品库存中提取信息，写入汇总 sheet ===
                # 先找生成的赛卓-成品库存 DataFrame
                product_inventory_pivoted = None
                for f in uploaded_files:
                    if f.name == "赛卓-成品库存.xlsx":
                        df_product_inventory = pd.read_excel(f)
                        config_inventory = CONFIG['pivot_config']['赛卓-成品库存.xlsx']
                        if 'date_format' in config_inventory and config_inventory['columns'] in df_product_inventory.columns:
                            df_product_inventory = process_date_column(df_product_inventory, config_inventory['columns'], config_inventory['date_format'])
                        product_inventory_pivoted = create_pivot(df_product_inventory, config_inventory, f.name, mapping_df)
                        break
                
                if product_inventory_pivoted is not None:
                    required_columns = ['WAFER品名', '规格', '品名', '数量_HOLD仓', '数量_成品仓', '数量_半成品仓']
                    if all(col in product_inventory_pivoted.columns for col in required_columns):
                        # 构造 key
                        inventory_key = product_inventory_pivoted[['WAFER品名', '规格', '品名']].astype(str)
                        summary_key = unfulfilled_orders_summary[['晶圆品名', '规格', '品名']].astype(str)
                
                        product_inventory_pivoted['已匹配'] = False
                
                        # 定位汇总 sheet
                        summary_sheet = writer.sheets['汇总']
                        start_col = summary_sheet.max_column + 1
                
                        # 合并第一行写“成品库存”
                        summary_sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col+2)
                        summary_sheet.cell(row=1, column=start_col, value='成品库存').alignment = Alignment(horizontal='center', vertical='center')
                        summary_sheet.cell(row=2, column=start_col, value='数量_HOLD仓').alignment = Alignment(horizontal='center', vertical='center')
                        summary_sheet.cell(row=2, column=start_col+1, value='数量_成品仓').alignment = Alignment(horizontal='center', vertical='center')
                        summary_sheet.cell(row=2, column=start_col+2, value='数量_半成品仓').alignment = Alignment(horizontal='center', vertical='center')
                
                        # 遍历汇总表行（从第3行开始）
                        for row_idx in range(3, summary_sheet.max_row + 1):
                            summary_wf = summary_sheet.cell(row=row_idx, column=1).value
                            summary_spec = summary_sheet.cell(row=row_idx, column=2).value
                            summary_prod = summary_sheet.cell(row=row_idx, column=3).value
                
                            match = product_inventory_pivoted[
                                (product_inventory_pivoted['WAFER品名'].astype(str) == str(summary_wf)) &
                                (product_inventory_pivoted['规格'].astype(str) == str(summary_spec)) &
                                (product_inventory_pivoted['品名'].astype(str) == str(summary_prod))
                            ]
                
                            if not match.empty:
                                hold = match['数量_HOLD仓'].values[0]
                                finished = match['数量_成品仓'].values[0]
                                semi_finished = match['数量_半成品仓'].values[0]
                                summary_sheet.cell(row=row_idx, column=start_col, value=hold)
                                summary_sheet.cell(row=row_idx, column=start_col+1, value=finished)
                                summary_sheet.cell(row=row_idx, column=start_col+2, value=semi_finished)
                
                                product_inventory_pivoted.loc[match.index, '已匹配'] = True
                
                        # 在赛卓-成品库存 sheet 中标红未匹配行
                        inventory_sheet = writer.book['赛卓-成品库存']
                        for row_idx, matched in enumerate(product_inventory_pivoted['已匹配'], start=2):
                            if not matched:
                                for col_idx in range(1, len(product_inventory_pivoted.columns) + 1):
                                    inventory_sheet.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                ### 成品在制
                semi_finished_value = 0
                semi_row = pd.DataFrame()  # ✅ 先定义为空 DataFrame

                product_in_progress_pivoted = None
                for f in uploaded_files:
                    if f.name == "赛卓-成品在制.xlsx":
                        df_product_in_progress = pd.read_excel(f)
                        config_in_progress = CONFIG['pivot_config']['赛卓-成品在制.xlsx']
                        if 'date_format' in config_in_progress and config_in_progress['columns'] in df_product_in_progress.columns:
                            df_product_in_progress = process_date_column(df_product_in_progress, config_in_progress['columns'], config_in_progress['date_format'])
                        product_in_progress_pivoted = create_pivot(df_product_in_progress, config_in_progress, f.name, mapping_df)
                        break
                
                if product_in_progress_pivoted is not None:
                    numeric_cols = product_in_progress_pivoted.select_dtypes(include='number').columns.tolist()
                    product_in_progress_pivoted['已匹配'] = False
                
                    summary_sheet = writer.sheets['汇总']
                    start_col = summary_sheet.max_column + 1
                
                    summary_sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 1)
                    summary_sheet.cell(row=1, column=start_col, value='赛卓-成品在制').alignment = Alignment(horizontal='center', vertical='center')
                    summary_sheet.cell(row=2, column=start_col, value='成品').alignment = Alignment(horizontal='center', vertical='center')
                    summary_sheet.cell(row=2, column=start_col + 1, value='半成品').alignment = Alignment(horizontal='center', vertical='center')
                
                    for row_idx in range(3, summary_sheet.max_row + 1):
                        summary_wf = summary_sheet.cell(row=row_idx, column=1).value
                        summary_spec = summary_sheet.cell(row=row_idx, column=2).value
                        summary_prod = summary_sheet.cell(row=row_idx, column=3).value
                
                        # 直接找成品
                        match = product_in_progress_pivoted[
                            (product_in_progress_pivoted['晶圆型号'].astype(str) == str(summary_wf)) &
                            (product_in_progress_pivoted['产品规格'].astype(str) == str(summary_spec)) &
                            (product_in_progress_pivoted['产品品名'].astype(str) == str(summary_prod))
                        ]
                
                        finished_value = match[numeric_cols].sum(axis=1).values[0] if not match.empty else 0

                        # 先找 mapping 表中满足晶圆品名、规格、品名、且半成品列非空的行
                        semi_match = df_full_mapping[
                            (df_full_mapping['新晶圆品名'].astype(str) == str(summary_wf)) &
                            (df_full_mapping['新规格'].astype(str) == str(summary_spec)) &
                            (df_full_mapping['半成品'].notnull()) &
                            (df_full_mapping['半成品'].astype(str) != '')
                        ]
                        
                        semi_finished_value = 0
                        
                        if not semi_match.empty:
                            semi_wafer = semi_match['新晶圆品名'].values[0]
                            semi_spec = semi_match['新规格'].values[0]
                            semi_prod = semi_match['新品名'].values[0]
                        
                            # 打印 mapping 表匹配到的半成品 key
                            st.write(f"✅ Mapping 匹配到半成品 → 晶圆型号: {semi_wafer}, 产品规格: {semi_spec}, 产品品名: {semi_prod}")
                        
                            semi_row = product_in_progress_pivoted[
                                (product_in_progress_pivoted['晶圆型号'].astype(str) == str(semi_wafer)) &
                                (product_in_progress_pivoted['产品规格'].astype(str) == str(semi_spec)) &
                                (product_in_progress_pivoted['产品品名'].astype(str) == str(semi_prod))
                            ]
                        
                            if not semi_row.empty:
                                semi_finished_value = semi_row[numeric_cols].sum(axis=1).values[0]
                        
                                # 打印成品 → 在成品在制里找到了半成品对应行的提示
                                st.write(f"🎯 成品 → 在成品在制里找到了半成品行 → 成品: {summary_wf}, {summary_spec}, {summary_prod} | 半成品: {semi_wafer}, {semi_spec}, {semi_prod}")
                            else:
                                semi_finished_value = 0
                        
                                # 打印成品 → 没找到半成品对应行的提示
                                st.write(f"⚠️ 成品 → 没在成品在制里找到半成品行 → 成品: {summary_wf}, {summary_spec}, {summary_prod} | 半成品: {semi_wafer}, {semi_spec}, {semi_prod}")

                        
                        # 写入到汇总表
                        summary_sheet.cell(row=row_idx, column=start_col, value=finished_value)
                        summary_sheet.cell(row=row_idx, column=start_col + 1, value=semi_finished_value)
                
                        # 标记成品匹配
                        if not match.empty:
                            product_in_progress_pivoted.loc[match.index, '已匹配'] = True
                        if not semi_row.empty:
                            product_in_progress_pivoted.loc[semi_row.index, '已匹配'] = True
                
                    # 在赛卓-成品在制 sheet 中标红未匹配行
                    in_progress_sheet = writer.book['赛卓-成品在制']
                    for row_idx, matched in enumerate(product_in_progress_pivoted['已匹配'], start=2):
                        if not matched:
                            for col_idx in range(1, len(product_in_progress_pivoted.columns) + 1):
                                in_progress_sheet.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")



                # 自动调整列宽
                for idx, col in enumerate(worksheet.columns, 1):
                    col_letter = get_column_letter(idx)
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                cell_len = sum(2 if ord(char) > 127 else 1 for char in str(cell.value))
                                max_length = max(max_length, cell_len)
                        except:
                               pass
                    worksheet.column_dimensions[col_letter].width = max_length + 5

            # === 在汇总 sheet 加黑框 ===
            summary_sheet = writer.book['汇总']
            # 假设前两行是标题
            max_row = 2  
            max_col = summary_sheet.max_column
            add_black_border(summary_sheet, max_row, max_col)

                    



        # 下载按钮
        with open(CONFIG['output_file'], 'rb') as f:
            st.download_button('下载汇总报告', f, CONFIG['output_file'])



if __name__ == '__main__':
    main()
